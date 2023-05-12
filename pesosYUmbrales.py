import numpy as np
import openpyxl

def guardarPesos(matriz,nombre):
  matriz= matriz.tolist()
  workbook = openpyxl.Workbook()
  sheet = workbook.active
  for row in matriz:
    sheet.append(row)
    rutaPesos="archivosExcelGuardados/"+nombre+'.xlsx'
    workbook.save(rutaPesos)
  return rutaPesos
  
def guardarUmbrales(matriz,nombre):
  matriz=matriz.tolist()
  workbook = openpyxl.Workbook()
  sheet = workbook.active 
  sheet.append(matriz)
  rutaUmbrales="archivosExcelGuardados/"+nombre+'.xlsx'
  workbook.save(rutaUmbrales)
  return rutaUmbrales

def matrizPesos(entradas, salidas):
    matrizPesos = np.random.uniform(low=-1, high=1, size=(salidas, entradas))
    matrizPesos = np.round(matrizPesos, decimals=2)
    return matrizPesos


def matrizUmbrales(salidas):
    matrizUmbrales = np.random.uniform(low=-1, high=1, size=(salidas))
    matrizUmbrales = np.round(matrizUmbrales, decimals=2)
    return matrizUmbrales


def readDataExcel(ruta):    
    # Define variable to load the dataframe
    excel_dataframe = openpyxl.load_workbook(ruta)
    # Define variable to read sheet
    dataframe = excel_dataframe.active
    print(dataframe)
    data = []
# Recorre el archibo de umbrales y los guarda en data y los imprime
    for col in dataframe.iter_cols(0, dataframe.max_column):
      for row in range(0, dataframe.max_row):       
        data.append(col[row].value)
    
    return data



def readDataExcel2(ruta):    
    # Define variable to load the dataframe
    excel_dataframe = openpyxl.load_workbook(ruta)
    # Define variable to read sheet
    dataframe = excel_dataframe.active
    print(dataframe)
    data = []
    data2 = []
# Recorre el archibo de umbrales y los guarda en data y los imprime
    # for col in dataframe.iter_cols(0, dataframe.max_column):
    for row in range(0, dataframe.max_row):
      data2=[]
      # for row in range(0, dataframe.max_row):       
      for col in dataframe.iter_cols(0, dataframe.max_column):       
        data2.append(col[row].value)
      data.append(data2) 
    return data


# entradas = int(input("Entradas: \n"))
# salidas = int(input("Salidas: \n"))
# matrizp=matrizPesos(entradas, salidas)
# print(matrizp)
# print("Matriz De umbrales")
# matrizu =matrizUmbrales(salidas)
# print(matrizu)
# rutaPesos= guardarPesos(matrizp,"pesos")
# rutaUnmbrales= guardarUmbrales(matrizu,"Umbrales")
# print(rutaPesos)
# lecturaMatrizP=readDataExcel(rutaPesos)
# print(lecturaMatrizP)


