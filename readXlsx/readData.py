import openpyxl
from tabulate import tabulate


def readData(ruta):

    # Define variable to load the dataframe
    excel_dataframe = openpyxl.load_workbook(ruta)

    # Define variable to read sheet
    dataframe = excel_dataframe.active

    print(dataframe)

    _rows = []
    dataEntraces = []
    dataExit = []

    # se definen las variables de entradas y salidas de la matriz
    entradas = 0
    salidas = 0
    valoresNulos = 0
    # recorre las columnas para contar entradas y salidas
    for col in dataframe.iter_cols(0, dataframe.max_column):

        if "X" in col[0].value:
            entradas += 1
        else:
            if "S" in col[0].value:
                salidas += 1
            else:
                valoresNulos += 1
    # variable de la cantidad de patrones
    patrones = 0
    # recorre las filas para contar los patrones
    for row in range(1, dataframe.max_row):
        patrones += 1
# Recorre el archivo sacando las entradas correspondientes y las almacena por patrones de entrada
    for row in range(1, dataframe.max_row):
        for col in dataframe.iter_cols(1, entradas):
            _rows.append(col[row].value)
            longitudDeRows = len(_rows)
            if longitudDeRows == entradas:
                dataEntraces.append(_rows)
                _rows = []
# Recorre el archivo sacando las salidas correspondientes y las almacena por patrones de salida
    for row in range(1, dataframe.max_row):
        for col in dataframe.iter_cols(entradas+1, dataframe.max_column):
            _rows.append(col[row].value)
            longitudDeRows = len(_rows)
            if longitudDeRows == salidas:
                dataExit.append(_rows)
                _rows = []

    return entradas, salidas, patrones, dataEntraces, dataExit

    # data.append(_row)

    # print(tabulate(data))

    # print(tabulate(data,headers=encabezado))
